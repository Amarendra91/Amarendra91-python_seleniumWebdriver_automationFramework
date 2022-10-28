import openpyxl


class HomePageTestData:
    date = [{"birthday": "31-12-1999"}]

    @staticmethod
    def getUserData(userinfo):
        Dict = {}
        workBook = openpyxl.load_workbook("C:\\Users\\amare\\PycharmProjects\\"
                                          "WebAutomationFrameworkDesign[Python,Selenium]""\\UserData.xlsx")
        workSheet = workBook.active
        for i in range(1, workSheet.max_row + 1):  # to get row info
            if workSheet.cell(row=i, column=1).value == userinfo:
                for j in range(2, workSheet.max_column + 1):  # to get column info
                    Dict[workSheet.cell(row=1, column=j).value] = workSheet.cell(row=i, column=j).value
        return [Dict]
